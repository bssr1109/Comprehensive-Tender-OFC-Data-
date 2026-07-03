import streamlit as st
import pandas as pd
from supabase import create_client
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

st.set_page_config(page_title="OFC Route Mapping", layout="wide")

SUPABASE_URL = st.secrets["SUPABASE_URL"]
SUPABASE_KEY = st.secrets["SUPABASE_KEY"]
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

MASTER_NAME = "seshu"


def get_routes(hq=None):
    q = supabase.table("ofc_routes_master").select("*").order("id")
    if hq:
        q = q.eq("headquarters", hq)
    return q.execute().data


def get_hqs():
    response = supabase.table("ofc_routes_master").select("headquarters").execute()
    data = response.data or []

    hqs = []
    for row in data:
        val = row.get("headquarters")
        if val and val.strip():
            hqs.append(val.strip())

    return sorted(list(set(hqs)))


def get_sections(route_id):
    return supabase.table("ofc_route_sections").select("*").eq("route_id", route_id).order("section_no").execute().data


def get_status(route_id):
    data = supabase.table("route_entry_status").select("*").eq("route_id", route_id).execute().data
    return data[0] if data else None


def save_route(route, sections, entered_by):
    route_id = route["id"]

    supabase.table("ofc_route_sections").delete().eq("route_id", route_id).execute()

    clean_sections = []
    for i, sec in enumerate(sections, start=1):
        clean_sections.append({
            "route_id": route_id,
            "section_no": i,
            "km_from": float(sec["km_from"]),
            "km_to": float(sec["km_to"]),
            "fiber_type": sec["fiber_type"]
        })

    if clean_sections:
        supabase.table("ofc_route_sections").insert(clean_sections).execute()

    last_end = float(clean_sections[-1]["km_to"]) if clean_sections else 0
    completed = round(last_end, 2) == round(float(route["total_rkm"]), 2)

    existing = get_status(route_id)
    payload = {
        "route_id": route_id,
        "entered_by": entered_by,
        "completed": completed
    }

    if existing:
        supabase.table("route_entry_status").update(payload).eq("route_id", route_id).execute()
    else:
        supabase.table("route_entry_status").insert(payload).execute()

    return completed


def make_excel():
    routes = get_routes()
    wb = Workbook()
    ws = wb.active
    ws.title = "OFC Final Report"

    headers = [
        "Sl.No.", "Route Name", "Transnet ID", "Route Length",
        "Route Capacity", "KM", "OH fiber", "UG fiber",
        "Entered By", "Timestamp"
    ]
    ws.append(headers)

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        cell.fill = header_fill

    row = 2
    sl = 1

    for route in routes:
        sections = get_sections(route["id"])
        status = get_status(route["id"])

        if not sections:
            sections = [{"km_from": "", "km_to": "", "fiber_type": ""}]

        start_row = row

        for sec in sections:
            ws.cell(row, 1, sl)
            ws.cell(row, 2, route.get("route_name", ""))
            ws.cell(row, 3, route.get("transnet_id", ""))
            ws.cell(row, 4, route.get("total_rkm", ""))
            ws.cell(row, 5, route.get("cable_size", ""))

            km_from = sec.get("km_from", "")
            km_to = sec.get("km_to", "")
            ws.cell(row, 6, f"{km_from} - {km_to}" if km_from != "" else "")

            if sec.get("fiber_type") == "OH":
                ws.cell(row, 7, "OH fiber")
            elif sec.get("fiber_type") == "UG":
                ws.cell(row, 8, "UG fiber")

            ws.cell(row, 9, status.get("entered_by", "") if status else "")
            ws.cell(row, 10, status.get("updated_at", "") if status else "")

            for col in range(1, 11):
                ws.cell(row, col).border = border
                ws.cell(row, col).alignment = Alignment(vertical="center", wrap_text=True)

            row += 1

        end_row = row - 1

        if end_row > start_row:
            for col in [1, 2, 3, 4, 5, 9, 10]:
                ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
                ws.cell(start_row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        sl += 1

    widths = [10, 45, 15, 15, 15, 15, 15, 15, 20, 25]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def login_page():
    st.title("OFC Route Physical Fiber Mapping")

    name = st.text_input("Enter Name")

    try:
        hqs = get_hqs()
    except Exception as e:
        st.error(f"Supabase read error: {e}")
        hqs = []

    if not hqs:
        st.warning("No headquarters found. Check Supabase RLS policy or master table data.")

    hq = st.selectbox("Select Headquarters", [""] + hqs)

    if st.button("Start"):
        if not name.strip():
            st.error("Enter name")
            return

        st.session_state.name = name.strip()

        if name.strip().lower() == MASTER_NAME:
            st.session_state.role = "master"
        else:
            if not hq:
                st.error("Select headquarters")
                return

            st.session_state.role = "field"
            st.session_state.hq = hq
            st.session_state.index = 0

        st.rerun()

def field_screen():
    name = st.session_state.name
    hq = st.session_state.hq

    routes = get_routes(hq)
    if not routes:
        st.warning("No routes found for this headquarters.")
        return

    idx = st.session_state.get("index", 0)
    idx = max(0, min(idx, len(routes) - 1))
    st.session_state.index = idx

    route = routes[idx]

    st.subheader(f"{hq}")
    st.write(f"Route {idx + 1} of {len(routes)}")

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Route Length", route["total_rkm"])
    c2.metric("Cable Size", route.get("cable_size", ""))
    c3.metric("Transnet ID", route.get("transnet_id", ""))
    c4.metric("User", name)

    st.markdown(f"### {route['route_name']}")

    existing_sections = get_sections(route["id"])

    if "sections" not in st.session_state or st.session_state.get("loaded_route") != route["id"]:
        if existing_sections:
            st.session_state.sections = [
                {
                    "km_from": float(x["km_from"]),
                    "km_to": float(x["km_to"]),
                    "fiber_type": x["fiber_type"]
                }
                for x in existing_sections
            ]
        else:
            st.session_state.sections = [{"km_from": 0.0, "km_to": 0.0, "fiber_type": "UG"}]

        st.session_state.loaded_route = route["id"]

    st.markdown("### Physical Fiber Sections")

    for i, sec in enumerate(st.session_state.sections):
        c1, c2, c3 = st.columns(3)

        sec["km_from"] = c1.number_input(
            f"KM From {i+1}",
            value=float(sec["km_from"]),
            step=0.01,
            key=f"from_{route['id']}_{i}"
        )

        sec["km_to"] = c2.number_input(
            f"KM To {i+1}",
            value=float(sec["km_to"]),
            step=0.01,
            key=f"to_{route['id']}_{i}"
        )

        sec["fiber_type"] = c3.selectbox(
            f"Type {i+1}",
            ["UG", "OH"],
            index=0 if sec["fiber_type"] == "UG" else 1,
            key=f"type_{route['id']}_{i}"
        )

    if st.button("+ Add Section"):
        last_to = st.session_state.sections[-1]["km_to"]
        st.session_state.sections.append({
            "km_from": float(last_to),
            "km_to": float(last_to),
            "fiber_type": "UG"
        })
        st.rerun()

    last_end = st.session_state.sections[-1]["km_to"]
    total_rkm = float(route["total_rkm"])

    if round(float(last_end), 2) == round(total_rkm, 2):
        st.success("Route length completed")
    else:
        st.warning(f"Current end KM: {last_end}. Route Length: {total_rkm}")

    b1, b2, b3, b4 = st.columns(4)

    if b1.button("Previous"):
        if idx > 0:
            st.session_state.index -= 1
            st.session_state.pop("sections", None)
            st.rerun()

    if b2.button("Save"):
        completed = save_route(route, st.session_state.sections, name)
        st.success("Saved successfully" + (" - Completed" if completed else " - Partial"))

    if b3.button("Save & Next"):
        save_route(route, st.session_state.sections, name)
        if idx < len(routes) - 1:
            st.session_state.index += 1
            st.session_state.pop("sections", None)
        st.rerun()

    if b4.button("Next"):
        if idx < len(routes) - 1:
            st.session_state.index += 1
            st.session_state.pop("sections", None)
            st.rerun()

    if st.button("Logout"):
        st.session_state.clear()
        st.rerun()


def master_dashboard():
    st.title("Master Dashboard")

    routes = get_routes()
    statuses = supabase.table("route_entry_status").select("*").execute().data

    total = len(routes)
    completed = sum(1 for x in statuses if x.get("completed"))
    partial = len(statuses) - completed
    pending = total - len(statuses)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total Routes", total)
    c2.metric("Completed", completed)
    c3.metric("Partial", partial)
    c4.metric("Pending", pending)

    st.subheader("HQ Wise Progress")

    df_routes = pd.DataFrame(routes)
    df_status = pd.DataFrame(statuses)

    if not df_routes.empty:
        for hq in sorted(df_routes["headquarters"].dropna().unique()):
            hq_routes = df_routes[df_routes["headquarters"] == hq]
            route_ids = set(hq_routes["id"])
            hq_completed = sum(
                1 for x in statuses
                if x["route_id"] in route_ids and x.get("completed")
            )
            st.write(f"**{hq}** — {hq_completed}/{len(hq_routes)} completed")
            st.progress(hq_completed / len(hq_routes))

    st.subheader("Download")
    excel_file = make_excel()

    st.download_button(
        "Download Final Excel Report",
        data=excel_file,
        file_name="OFC_Final_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.subheader("All Routes Status")

    rows = []
    status_map = {x["route_id"]: x for x in statuses}

    for r in routes:
        s = status_map.get(r["id"], {})
        rows.append({
            "HQ": r.get("headquarters"),
            "Route Name": r.get("route_name"),
            "Transnet ID": r.get("transnet_id"),
            "Total RKM": r.get("total_rkm"),
            "Cable Size": r.get("cable_size"),
            "Entered By": s.get("entered_by", ""),
            "Completed": s.get("completed", False),
            "Updated At": s.get("updated_at", "")
        })

    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    if st.button("Logout"):
        st.session_state.clear()
        st.rerun()


if "role" not in st.session_state:
    login_page()
else:
    if st.session_state.role == "master":
        master_dashboard()
    else:
        field_screen()